import pandas as pd
import numpy as np
import os
from openpyxl.utils import get_column_letter

def process_excel(filename, expected_columns):
    if not os.path.exists(filename):
        print(f"--> File {filename} tidak ditemukan.")
        return

    df_raw = pd.read_excel(filename, header=None)
    
    col_indices = {}
    max_header_row = -1

    for r_idx, row in df_raw.head(20).iterrows():
        for c_idx, val in row.items():
            if pd.notna(val):
                clean_val = str(val).strip()
                if clean_val in expected_columns and clean_val not in col_indices:
                    col_indices[clean_val] = c_idx
                    max_header_row = max(max_header_row, r_idx)

    missing_cols = [col for col in expected_columns if col not in col_indices]
    if missing_cols:
        print(f"--> Header {missing_cols} tidak ditemukan pada file {filename}.")
        return

    df_data = df_raw.iloc[max_header_row + 1:].copy()
    
    df_filtered = pd.DataFrame()
    for col in expected_columns:
        df_filtered[col] = df_data[col_indices[col]]

    df_filtered = df_filtered.replace(r'^\s*$', np.nan, regex=True)
    df_filtered = df_filtered.dropna(how='any')

    cols_to_format = ["Kode", "Qty", "SIZE", "DPP", "Jumlah"]

    def clean_number(x):
        if pd.isna(x):
            return x
        if isinstance(x, (int, float)):
            return float(x)
        xs = str(x).strip()
        if xs == '':
            return np.nan
        xs = xs.replace('.', '').replace(',', '.')
        try:
            return float(xs)
        except ValueError:
            return x

    for col in cols_to_format:
        if col in df_filtered.columns:
            df_filtered[col] = df_filtered[col].apply(clean_number)

    base_name, _ = os.path.splitext(filename)
    output_filename = f"{base_name}_temp.xlsx"

    writer = pd.ExcelWriter(output_filename, engine='openpyxl')
    df_filtered.to_excel(writer, index=False, sheet_name='Sheet1')
    
    worksheet = writer.sheets['Sheet1']
    for idx, col in enumerate(df_filtered.columns):
        max_len = len(str(col))
        col_letter = get_column_letter(idx + 1)
        
        if not df_filtered[col].empty:
            col_max = df_filtered[col].astype(str).str.len().max()
            if pd.notna(col_max):
                max_len = max(max_len, int(col_max))
        
        worksheet.column_dimensions[col_letter].width = max_len + 2
        
        if col in cols_to_format:
            for row_idx in range(2, worksheet.max_row + 1):
                worksheet[f"{col_letter}{row_idx}"].number_format = '#,##0.00'

    writer.close()
    print(f"--> Proses selesai, hasil disimpan di {output_filename}")

cols_ptmus = ["Gudang", "No. Pelanggan", "Nama Dept.", "No. Faktur", "Tgl Faktur", "Kode", "Nama Barang", "Qty", "Unit 1", "SIZE", "DPP", "Kategori"]
cols_ptmsu = ["Gudang", "No. Pelanggan", "Nama Dept.", "No. Faktur", "Tgl Faktur", "Kode", "Nama Barang", "Qty", "Unit 1", "SIZE", "Jumlah", "Kategori"]

process_excel("PTMUS.xls", cols_ptmus)
process_excel("PTMSU.xls", cols_ptmsu)
